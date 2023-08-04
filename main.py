from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import shutil
import os
import pandas as pd
from openpyxl import load_workbook
import PyPDF2
from PyPDF2 import PdfReader
import openpyxl





driver = webdriver.Chrome()

driver.get("https://dejt.jt.jus.br/dejt/f/n/diariocon")

def data_inicio():
    driver.find_element(By.XPATH, "/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[1]/fieldset/table/tbody/tr/td[1]/table/tbody/tr/td[1]/span/input").clear()
    driver.find_element(By.XPATH, "/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[1]/fieldset/table/tbody/tr/td[1]/table/tbody/tr/td[1]/span/input").send_keys("26/07/2023")

def data_fim():
    driver.find_element(By.XPATH, "/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[1]/fieldset/table/tbody/tr/td[2]/table/tbody/tr/td[1]/span/input").clear()
    driver.find_element(By.XPATH, "/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[1]/fieldset/table/tbody/tr/td[2]/table/tbody/tr/td[1]/span/input").send_keys("02/08/2023")


def orgao():
    tst = driver.find_element(By.XPATH, '/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[1]/fieldset/table/tbody/tr/td[4]/span[2]/span/select/option[2]')
    tst.click()

def pesquisar():
    pesquisar = driver.find_element(By.XPATH, '/html/body/div[1]/div[4]/form/span[1]/span[2]/div[3]/div[1]/button[3]')
    pesquisar.click()


def baixar_todos():
    driver.find_element(By.XPATH, '/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[3]/fieldset/table/tbody/tr[2]/td[3]/button').click()
    driver.find_element(By.XPATH, '/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[3]/fieldset/table/tbody/tr[3]/td[3]/button').click()
    driver.find_element(By.XPATH, '/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[3]/fieldset/table/tbody/tr[4]/td[3]/button').click()
    driver.find_element(By.XPATH, '/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[3]/fieldset/table/tbody/tr[6]/td[3]/button').click()
    driver.find_element(By.XPATH, '/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[3]/fieldset/table/tbody/tr[7]/td[3]/button').click()
    driver.find_element(By.XPATH, '/html/body/div[1]/div[4]/form/span[1]/span[3]/div[1]/div[3]/fieldset/table/tbody/tr[8]/td[3]/button').click()

data_inicio()
time.sleep(2)
data_fim()
time.sleep(2)
orgao()
time.sleep(2)
pesquisar()
time.sleep(2)
baixar_todos()
time.sleep(10)
driver.quit()


def mover_arquivo(origem, destino):
    try:
        # Verifica se o arquivo de origem existe
        if not os.path.isfile(origem):
            print(f"Arquivo de origem '{origem}' não encontrado.")
            return

        # Verifica se o destino é uma pasta existente
        if not os.path.isdir(destino):
            print(f"Pasta de destino '{destino}' não encontrada.")
            return

        # Move o arquivo para o destino
        shutil.move(origem, destino)
        print(f"Arquivo movido com sucesso de '{origem}' para '{destino}'.")
    except Exception as e:
        print(f"Ocorreu um erro ao mover o arquivo: {e}")



# # Exemplo de uso
# origem_arquivo = '../../../Downloads/Diario_3777__1_8_2023.pdf'
# destino_pasta = './arquivos/'

# mover_arquivo(origem_arquivo, destino_pasta)
import os

def find_downloads_folder():
    user_home = os.path.expanduser("~")  # Caminho para a pasta do usuário (diretório home)
    downloads_folder = os.path.join(user_home, "Downloads")  # Combinar o caminho com "Downloads"
    
    if os.path.exists(downloads_folder):  # Verificar se a pasta "Downloads" existe
        return downloads_folder
    else:
        print("Pasta Downloads não encontrada.")
        return None
    
# diretorio_atual = '../../../Downloads/'

# downloads_folder_path = find_downloads_folder()
diretorio_atual = find_downloads_folder()

# if downloads_folder_path:
#     print(f"Caminho para a pasta Downloads: {downloads_folder_path}")

nome_selecionado = "Diario"  # Substitua "exemplo" pelo nome que você deseja selecionar

# Caminho de destino para a pasta onde os arquivos serão movidos
destino_pasta = './diarios_baixados/'

# Listar arquivos no diretório atual (ou substitua o caminho do diretório)

arquivos_no_diretorio = os.listdir(diretorio_atual)

# # Loop para mover os arquivos selecionados
for nome_arquivo in arquivos_no_diretorio:
    if nome_arquivo.endswith('.pdf'):
        if nome_selecionado in nome_arquivo:
            caminho_origem = os.path.join(diretorio_atual, nome_arquivo)
            mover_arquivo(caminho_origem, destino_pasta)


############################################################################################################################################################################################



# Função para converter PDF para TXT
def converter_pdf_para_txt(caminho_arquivo_pdf, caminho_arquivo_txt):
    leitor = PdfReader(caminho_arquivo_pdf)
    texto = ''
    for pagina in leitor.pages:
        texto += pagina.extract_text()

    with open(caminho_arquivo_txt, 'w', encoding='utf-8') as arquivo_txt:
        arquivo_txt.write(texto)

# Diretório onde estão os arquivos PDF
pasta_pdf = './diarios_baixados/'

# Diretório onde serão salvos os arquivos TXT convertidos
pasta_txt = './diarios_txt/'

# Loop pelos arquivos na lista de arquivos PDF
for nome_arquivo_pdf in os.listdir(pasta_pdf):
    caminho_arquivo_pdf = os.path.join(pasta_pdf, nome_arquivo_pdf)

    # Defina o nome do arquivo TXT com a mesma base do PDF, mas com extensão .txt
    nome_arquivo_txt = os.path.splitext(nome_arquivo_pdf)[0] + '.txt'
    caminho_arquivo_txt = os.path.join(pasta_txt, nome_arquivo_txt)
    if nome_arquivo_pdf.endswith('.pdf'):
        converter_pdf_para_txt(caminho_arquivo_pdf, caminho_arquivo_txt)
        print(f'Arquivo {nome_arquivo_pdf} convertido para {nome_arquivo_txt}.')


########################################################################################################################################################################################################        


# Função para ler o conteúdo de um arquivo PDF
def ler_arquivo_txt(caminho_arquivo_txt):
    with open(caminho_arquivo_txt, 'r', encoding='utf-8') as arquivo_txt:
        conteudo = arquivo_txt.read()
    return conteudo

# Diretório onde estão os arquivos TXT

# Diretório onde estão os arquivos PDF
pasta_txt = './diarios_txt/'

def inserir_dados(planilha, dados):
    for linha in dados:
        planilha.append(linha)

# Loop pelos arquivos na pasta
from joblib import Parallel, delayed


# for nome_arquivo in os.listdir(pasta_txt):
def loop_arquivos(nome_arquivo_txt:str):
    
    # for nome_arquivo_txt in os.listdir(pasta_txt):
        if nome_arquivo_txt.endswith('.txt'):
            caminho_arquivo_txt = os.path.join(pasta_txt, nome_arquivo_txt)
            conteudo_txt = ler_arquivo_txt(caminho_arquivo_txt)
    
        # PROCESSO Nº TST-AIRR - 1001243-58.2021.5.02.0320
        caracteres_procurados = 'PROCESSO Nº'
        # Procurando caracteres específicos na string

        indices_ocorrencia = []

        for indice, _ in enumerate(conteudo_txt):
            print('linhas percorridas :', indice, "linhas a percorrer:",len(conteudo_txt))
            if conteudo_txt[indice:].startswith(caracteres_procurados):
                indicesomado = indice + 48
                if conteudo_txt[indice:indicesomado] not in indices_ocorrencia:
                    indices_ocorrencia.append(conteudo_txt[indice:indicesomado])
                    arquivo_excel = "tst_"+str(nome_arquivo_txt[13:25]).strip('.txt')
                    arquivo_excel += '.xlsx'
                    if not os.path.exists(arquivo_excel):
                        # Se o arquivo não existir, criar um novo arquivo Excel
                        novo_arquivo_excel = openpyxl.Workbook()
                    else:
                        # Se o arquivo já existir, abrir o arquivo existente
                        novo_arquivo_excel = openpyxl.load_workbook(arquivo_excel)

                    # Selecionar a planilha 'Dados' ou criar se não existir
                    planilha_dados = novo_arquivo_excel.get_sheet_by_name('Dados') if 'Dados' in novo_arquivo_excel.sheetnames else novo_arquivo_excel.create_sheet(title='Dados')
                    dados_planilha = [[conteudo_txt[indice:indicesomado]]]
                    
                    # Inserir dados na planilha 'Dados'
                    inserir_dados(planilha_dados, dados_planilha)
                    
                    # Salvar o arquivo Excel
                    novo_arquivo_excel.save(arquivo_excel)
                                    # input()
                elif conteudo_txt[indice:indicesomado] in indices_ocorrencia:
                    print("ja foi posto na planilha")
                    # adiciona uma informação ao texto original
                    with open("log_processos_repitidos.txt", "a") as arquivo:
                        arquivo.write(conteudo_txt[indice:indicesomado])

# Executando a função em paralelo para cada número na lista
Parallel(n_jobs=-1)(delayed(loop_arquivos)(nome_arquivo) for nome_arquivo in os.listdir(pasta_txt))